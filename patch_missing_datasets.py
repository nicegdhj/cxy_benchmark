#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
patch_missing_datasets.py

将 fmt_exp0316 中缺失的三个数据集结果补充到已生成的 aggregated_reports 目录：
  - telequad_gen_0_shot    → TeleQuAD
  - tele_exam_gen_0_shot_str → 通信工程师中级考试真题-主观题
  - teledata_gen_0_shot    → Tele-Data

用法：
    python patch_missing_datasets.py \
        --fmt-dir ~/Desktop/fmt_exp0316 \
        --target-dir ./outputs/aggregated_reports_20260319_133458 \
        --output-base-dir ./outputs
"""

import argparse
import glob
import json
import os
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

TARGET_TASKS = [
    "telequad_gen_0_shot",
    "tele_exam_gen_0_shot_str",
    "teledata_gen_0_shot",
]


def load_mappings(output_base_dir: str):
    task_mapping = {}
    task_mapping_file = os.path.join(output_base_dir, "评测任务文件名对应.xlsx")
    if os.path.exists(task_mapping_file):
        df = pd.read_excel(task_mapping_file)
        if len(df.columns) >= 2:
            for _, row in df.iterrows():
                k = str(row.iloc[0]).strip()
                v = str(row.iloc[1]).strip()
                if pd.notna(k) and pd.notna(v) and k != "nan" and v != "nan":
                    task_mapping[k] = v

    id_to_name = {}
    exp_mapping_file = os.path.join(output_base_dir, "实验设置.xlsx")
    if os.path.exists(exp_mapping_file):
        df = pd.read_excel(exp_mapping_file)
        id_col = "编号" if "编号" in df.columns else df.columns[0]
        df = df.dropna(subset=[id_col])
        set_col = df.columns[1]
        for _, row in df.iterrows():
            try:
                row_id = int(float(row[id_col]))
                set_name = str(row[set_col]).strip()
                if set_name.endswith(".json"):
                    set_name = set_name[:-5]
                id_to_name[row_id] = set_name
            except (ValueError, TypeError):
                continue
    return task_mapping, id_to_name


def get_mapped_suite(raw_suite: str, task_mapping: dict) -> str:
    if raw_suite in task_mapping:
        return task_mapping[raw_suite]
    stripped = raw_suite.removesuffix("_suite")
    return task_mapping.get(stripped, raw_suite)


def get_mapped_exp_name(raw_name: str, id_to_name: dict) -> str:
    if raw_name in ("baseline", "baseline_nothink"):
        return raw_name
    if raw_name.startswith("pt"):
        match = re.match(r"pt(\d+)_sf[t]?0(.*)", raw_name)
        if match:
            pt_id = int(match.group(1))
            extra = match.group(2)
            if pt_id in id_to_name:
                return id_to_name[pt_id] + extra
        match2 = re.search(r"pt(\d+)_", raw_name)
        if match2:
            pt_id = int(match2.group(1))
            if pt_id in id_to_name:
                return id_to_name[pt_id]
    return raw_name


def find_detail_jsonl_files(group_dir: Path, task_name: str) -> list[str]:
    """通过 infer_meta.json 找 task 对应的 detail JSONL 文件。"""
    infer_meta_path = group_dir / "infer_meta.json"
    if infer_meta_path.exists():
        meta = json.loads(infer_meta_path.read_text(encoding="utf-8"))
        task_info = meta.get("tasks", {}).get(task_name)
        if task_info:
            timestamp = task_info.get("timestamp")
            if timestamp:
                results_dir = group_dir / "details" / timestamp / "results"
                if results_dir.exists():
                    files = glob.glob(str(results_dir / "**" / "*_details.jsonl"), recursive=True)
                    if files:
                        return files
    # fallback：按 task 名前缀搜索
    prefix = task_name.split("_")[0]
    return glob.glob(str(group_dir / "details" / "**" / f"*{prefix}*_details.jsonl"), recursive=True)


def jsonl_to_rows(jsonl_path: str) -> list[dict]:
    rows = []
    with open(jsonl_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
                eval_res = obj.get("eval_res")
                eval_details = obj.get("eval_details")
                pred_raw = obj.get("prediction")
                origin_prompt = prediction_val = gold = None

                if isinstance(pred_raw, dict):
                    origin_prompt = pred_raw.get("origin_prompt", "null")
                    prediction_val = pred_raw.get("prediction", "null")
                    gold = pred_raw.get("gold", "null")
                elif isinstance(pred_raw, str):
                    try:
                        parsed = json.loads(pred_raw)
                        if isinstance(parsed, dict):
                            origin_prompt = parsed.get("origin_prompt", "null")
                            prediction_val = parsed.get("prediction", "null")
                            gold = parsed.get("gold", "null")
                        else:
                            prediction_val = pred_raw
                    except json.JSONDecodeError:
                        prediction_val = pred_raw

                def fmt(v):
                    if v is None:
                        return "null"
                    if isinstance(v, (dict, list)):
                        return json.dumps(v, ensure_ascii=False, indent=2)
                    return str(v)

                rows.append({
                    "eval_res": fmt(eval_res),
                    "eval_details": fmt(eval_details),
                    "origin_prompt": fmt(origin_prompt),
                    "prediction": fmt(prediction_val),
                    "gold": fmt(gold),
                })
            except json.JSONDecodeError:
                continue
    return rows


def patch(fmt_dir: str, target_dir: str, output_base_dir: str):
    fmt_path = Path(fmt_dir).expanduser()
    target_path = Path(target_dir)

    task_mapping, id_to_name = load_mappings(output_base_dir)

    # ── 1. 扫描所有实验组，读取 3 个目标任务的 accuracy ──────────────────────
    # raw_name → {task_name: accuracy}
    group_accuracies: dict[str, dict] = {}
    # raw_name → mapped_name
    group_mapped: dict[str, str] = {}

    for entry in sorted(fmt_path.iterdir()):
        if not entry.is_dir():
            continue
        raw_name = entry.name
        report_path = entry / "report.json"
        if not report_path.exists():
            print(f"⚠️  跳过 {raw_name}：无 report.json")
            continue
        data = json.loads(report_path.read_text(encoding="utf-8"))
        accs = {}
        for t in data.get("tasks", []):
            tname = t.get("suite", t.get("task", ""))
            if tname in TARGET_TASKS:
                accs[tname] = t.get("accuracy")
        if not accs:
            print(f"⚠️  跳过 {raw_name}：未找到目标任务")
            continue
        group_accuracies[raw_name] = accs
        group_mapped[raw_name] = get_mapped_exp_name(raw_name, id_to_name)
        print(f"✅ {raw_name} → {group_mapped[raw_name]}：{accs}")

    if not group_accuracies:
        print("❌ 未找到任何实验组包含目标任务，退出")
        return

    # ── 2. 更新总体汇总 Excel ─────────────────────────────────────────────────
    excel_files = list(target_path.glob("总体汇总_*.xlsx"))
    if not excel_files:
        print("❌ 未找到总体汇总 Excel")
        return
    excel_path = excel_files[0]
    print(f"\n📊 更新 Excel：{excel_path}")

    wb = load_workbook(excel_path)
    ws = wb["总体对比"]

    # 读取列头（第1行），建立 mapped_exp_name → 列号 的映射
    col_count = ws.max_column
    header = {ws.cell(1, c).value: c for c in range(2, col_count + 1)}

    # 找附录起始行（插入新行的位置）
    appendix_row = None
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val and "附录" in str(val):
            appendix_row = r
            break
    # 数据行结束：appendix_row - 1（可能有空行）；找最后一个有数值的行
    last_data_row = 1
    for r in range(1, appendix_row or ws.max_row + 1):
        if isinstance(ws.cell(r, 2).value, (int, float)):
            last_data_row = r

    insert_at = last_data_row + 1  # 在最后一条数据行后插入

    # 插入 3 行（将附录等内容整体下移）
    ws.insert_rows(insert_at, len(TARGET_TASKS))
    print(f"  在第 {insert_at} 行插入 {len(TARGET_TASKS)} 行新数据")

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill  = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
    red_fill    = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    baseline_col = header.get("baseline", 2)

    for i, task_name in enumerate(TARGET_TASKS):
        row = insert_at + i
        mapped_task = get_mapped_suite(task_name, task_mapping)
        ws.cell(row, 1).value = mapped_task

        # baseline 值（用于着色参考）
        baseline_val = None
        if "baseline" in group_accuracies and task_name in group_accuracies["baseline"]:
            baseline_val = group_accuracies["baseline"][task_name]

        for raw_name, accs in group_accuracies.items():
            mapped_exp = group_mapped[raw_name]
            col = header.get(mapped_exp)
            if col is None:
                continue
            acc = accs.get(task_name)
            if acc is None:
                continue
            cell = ws.cell(row, col)
            cell.value = acc
            # 着色
            if col == baseline_col:
                cell.fill = yellow_fill
            elif baseline_val is not None and isinstance(acc, (int, float)):
                cell.fill = green_fill if acc > baseline_val else red_fill

        print(f"  ✓ 写入任务行：{mapped_task}")

    wb.save(excel_path)
    print(f"  💾 已保存 Excel")

    # ── 3. 创建明细子目录并生成 Excel ────────────────────────────────────────
    for raw_name, accs in group_accuracies.items():
        mapped_exp = group_mapped[raw_name]
        group_dir = fmt_path / raw_name

        for task_name in TARGET_TASKS:
            if task_name not in accs:
                continue
            mapped_task = get_mapped_suite(task_name, task_mapping)
            task_out_dir = target_path / mapped_task / mapped_exp
            task_out_dir.mkdir(parents=True, exist_ok=True)

            jsonl_files = find_detail_jsonl_files(group_dir, task_name)
            if not jsonl_files:
                print(f"  ⚠️  {raw_name}/{task_name}：未找到 detail JSONL")
                continue

            for jf in jsonl_files:
                rows = jsonl_to_rows(jf)
                if rows:
                    out_name = os.path.splitext(os.path.basename(jf))[0] + ".xlsx"
                    out_path = task_out_dir / out_name
                    pd.DataFrame(rows).to_excel(str(out_path), index=False)
                    print(f"  📄 {out_path}")


def main():
    parser = argparse.ArgumentParser(description="补充缺失数据集到已生成的 aggregated_reports")
    parser.add_argument("--fmt-dir", required=True, help="fmt 根目录")
    parser.add_argument("--target-dir", required=True, help="待补充的 aggregated_reports 目录")
    parser.add_argument("--output-base-dir", default="./outputs", help="outputs 根目录（含映射 Excel）")
    args = parser.parse_args()

    patch(
        fmt_dir=args.fmt_dir,
        target_dir=args.target_dir,
        output_base_dir=args.output_base_dir,
    )
    print("\n✅ 补充完成")


if __name__ == "__main__":
    main()
