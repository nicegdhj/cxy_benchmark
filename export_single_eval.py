#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
export_single_eval.py - 将单次评测结果（mixed_eval_*/eval_init/summary/）导出为 Excel

用法：
    python export_single_eval.py \\
        --eval-dir /path/to/mixed_eval_20260330_161500 \\
        --eval-version eval_init \\
        --output-dir /path/to/benchmark/outputs

    # 不传 --output-dir 则输出到 eval-dir/../benchmark/outputs/
"""

import argparse
import glob
import json
import os
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment


# 使用与 aggregate_eval_reports.py 相同的映射逻辑
def load_task_mapping(output_base_dir: str):
    task_mapping = {}
    task_mapping_file = os.path.join(output_base_dir, "评测任务文件名对应.xlsx")
    if os.path.exists(task_mapping_file):
        df = pd.read_excel(task_mapping_file)
        if len(df.columns) >= 2:
            for _, row in df.iterrows():
                k = str(row.iloc[0]).strip()
                v = str(row.iloc[1]).strip()
                if pd.notna(k) and pd.notna(v) and k != "nan" and v != "nan":
                    task_mapping[k] = v
    return task_mapping


def get_mapped_suite(raw_suite: str, task_mapping: dict) -> str:
    if raw_suite in task_mapping:
        return task_mapping[raw_suite]
    stripped = raw_suite.removesuffix("_suite")
    return task_mapping.get(stripped, raw_suite)


def read_summary_csv(csv_path: str) -> pd.DataFrame:
    try:
        df = pd.read_csv(csv_path)
        return df
    except Exception:
        return pd.DataFrame()


def _try_float(val) -> float | None:
    if val is None or val == "-":
        return None
    if isinstance(val, float) and pd.isna(val):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def extract_main_accuracy(df: pd.DataFrame, suite_name: str) -> float | None:
    """
    从 summary CSV 中提取主 accuracy（兼容 OpenCompass 多种汇总格式）。
    优先级：
    1. version=='-' 的汇总行，依次尝试 accuracy / score / naive_average / weighted_average / 任意指标
    2. 无汇总行时，对所有有效数值行取平均
    """
    if df.empty or len(df.columns) < 5:
        return None

    model_col = df.columns[-1]

    def get_val(row):
        return _try_float(row[model_col])

    if "version" not in df.columns or "metric" not in df.columns:
        # 无标准列，取第一个有效值
        for _, row in df.iterrows():
            v = get_val(row)
            if v is not None:
                return v
        return None

    # 汇总行（version == '-'）
    agg = df[df["version"] == "-"]
    if not agg.empty:
        for preferred in ("accuracy", "score", "naive_average", "weighted_average"):
            rows = agg[agg["metric"] == preferred]
            if not rows.empty:
                v = get_val(rows.iloc[0])
                if v is not None:
                    return v
        # 任意指标的汇总行
        for _, row in agg.iterrows():
            v = get_val(row)
            if v is not None:
                return v

    # 无汇总行：对所有有效数值行取平均
    valid_vals = [get_val(row) for _, row in df.iterrows() if get_val(row) is not None]
    if valid_vals:
        return round(sum(valid_vals) / len(valid_vals), 2)

    return None


def extract_suite_metrics(df: pd.DataFrame) -> list[dict]:
    """提取所有指标行（用于详情 sheet）。"""
    if df.empty:
        return []
    rows = []
    model_col = df.columns[-1] if len(df.columns) >= 1 else None
    for _, row in df.iterrows():
        val = row.get(model_col) if model_col else None
        if val == "-" or (isinstance(val, float) and pd.isna(val)):
            val = None
        else:
            try:
                val = float(val)
            except (ValueError, TypeError):
                val = None
        rows.append({
            "dataset": str(row.get("dataset", "")),
            "metric": str(row.get("metric", "")),
            "value": val,
        })
    return rows


def process(eval_dir: str, eval_version: str, output_base_dir: str):
    eval_dir = Path(eval_dir).resolve()
    summary_dir = eval_dir / eval_version / "summary"
    results_dir = eval_dir / eval_version / "results"
    meta_file = eval_dir / "infer_meta.json"

    if not summary_dir.exists():
        print(f"❌ 找不到 summary 目录：{summary_dir}")
        sys.exit(1)

    # 读取 meta
    meta = {}
    model_name = "model"
    if meta_file.exists():
        meta = json.loads(meta_file.read_text(encoding="utf-8"))
        model_name = meta.get("model_name", "model")

    task_mapping = load_task_mapping(output_base_dir)

    # 时间戳
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target_dir = os.path.join(output_base_dir, f"{eval_dir.name}_report_{timestamp}")
    os.makedirs(target_dir, exist_ok=True)
    out_excel = os.path.join(target_dir, f"{eval_dir.name}_汇总.xlsx")

    # ── 扫描所有 suite ────────────────────────────────────────────────────────
    suite_dirs = sorted([d for d in summary_dir.iterdir() if d.is_dir()])

    overview_rows = []
    suite_details = {}  # suite_name -> list of metric dicts

    for suite_dir in suite_dirs:
        suite_name = suite_dir.name
        csv_files = sorted(glob.glob(str(suite_dir / "*.csv")))
        if not csv_files:
            print(f"⚠️  {suite_name}：无 CSV 文件，跳过")
            overview_rows.append({
                "套件 (raw)": suite_name,
                "任务名": get_mapped_suite(suite_name, task_mapping),
                "类型": meta.get("tasks", {}).get(suite_name, {}).get("type", "-"),
                "准确率 (%)": None,
                "样本数": meta.get("tasks", {}).get(suite_name, {}).get("num_samples"),
                "状态": "no_csv",
            })
            continue

        # 取最新的 csv（按文件名排序最后一个）
        csv_path = csv_files[-1]
        df = read_summary_csv(csv_path)
        acc = extract_main_accuracy(df, suite_name)
        metrics = extract_suite_metrics(df)

        task_meta = meta.get("tasks", {}).get(suite_name, {})
        overview_rows.append({
            "套件 (raw)": suite_name,
            "任务名": get_mapped_suite(suite_name, task_mapping),
            "类型": task_meta.get("type", "-"),
            "准确率 (%)": acc,
            "样本数": task_meta.get("num_samples"),
            "状态": task_meta.get("status", "unknown") if task_meta else "unknown",
        })
        suite_details[suite_name] = metrics
        status = "✅" if acc is not None else "⚠️ (无有效分数)"
        print(f"  {status} {suite_name}: {acc}")

    df_overview = pd.DataFrame(overview_rows)

    # ── 写 Excel ─────────────────────────────────────────────────────────────
    with pd.ExcelWriter(out_excel, engine="openpyxl") as writer:
        df_overview.to_excel(writer, sheet_name="总体概览", index=False)
        ws = writer.sheets["总体概览"]

        # 格式：准确率列数值着色
        yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        bold = Font(bold=True)

        # 标题行加粗
        for cell in ws[1]:
            cell.font = bold

        # 准确率列 = D列(4)
        for row in range(2, len(df_overview) + 2):
            val_cell = ws.cell(row=row, column=4)
            status_cell = ws.cell(row=row, column=6)
            val = val_cell.value
            if val is None:
                val_cell.fill = gray
            elif isinstance(val, (int, float)):
                if val >= 70:
                    val_cell.fill = green
                elif val >= 50:
                    val_cell.fill = yellow
                else:
                    val_cell.fill = red
            if status_cell.value in ("error", "failed"):
                status_cell.fill = red

        # 调整列宽
        col_widths = [25, 25, 10, 15, 10, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        # 各 suite 详情 sheet
        for suite_name, metrics in suite_details.items():
            if not metrics:
                continue
            df_det = pd.DataFrame(metrics)
            sheet_name = suite_name[:31]
            base = sheet_name
            idx = 1
            while sheet_name in writer.sheets:
                sheet_name = f"{base[:28]}_{idx}"
                idx += 1
            df_det.to_excel(writer, sheet_name=sheet_name, index=False)

        # infer_meta sheet
        if meta:
            tasks_meta_rows = []
            for k, v in meta.get("tasks", {}).items():
                tasks_meta_rows.append({"suite": k, **v})
            if tasks_meta_rows:
                pd.DataFrame(tasks_meta_rows).to_excel(writer, sheet_name="infer_meta", index=False)

    print(f"\n📊 Excel 已输出：{out_excel}")

    # ── 明细 JSONL → Excel（与 aggregate_eval_reports.py 相同格式）────────────
    # 输出路径：target_dir/{mapped_suite}/{model_name}/{filename}.xlsx
    if results_dir.exists():
        print("\n📂 处理明细 JSONL...")
        for suite_name in suite_details.keys():
            mapped_suite = get_mapped_suite(suite_name, task_mapping)
            suite_results_dir = results_dir / suite_name
            if not suite_results_dir.exists():
                suite_results_dir = results_dir  # 兼容扁平结构

            jsonl_files = glob.glob(str(suite_results_dir / "**" / "*.jsonl"), recursive=True)
            if not jsonl_files:
                jsonl_files = glob.glob(
                    str(results_dir / "**" / f"*{suite_name}*_details.jsonl"), recursive=True
                )
            if not jsonl_files:
                continue

            task_out_dir = os.path.join(target_dir, mapped_suite, model_name)
            os.makedirs(task_out_dir, exist_ok=True)

            for jf in jsonl_files:
                rows = []
                with open(jf, encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        if not line:
                            continue
                        try:
                            obj = json.loads(line)
                            eval_res = obj.get("eval_res")
                            eval_details = obj.get("eval_details")
                            pred_raw = obj.get("prediction")
                            origin_prompt = prediction_val = gold = None

                            if isinstance(pred_raw, dict):
                                origin_prompt = pred_raw.get("origin_prompt", "null")
                                prediction_val = pred_raw.get("prediction", "null")
                                gold = pred_raw.get("gold", "null")
                            elif isinstance(pred_raw, str):
                                try:
                                    parsed = json.loads(pred_raw)
                                    if isinstance(parsed, dict):
                                        origin_prompt = parsed.get("origin_prompt", "null")
                                        prediction_val = parsed.get("prediction", "null")
                                        gold = parsed.get("gold", "null")
                                    else:
                                        prediction_val = pred_raw
                                except json.JSONDecodeError:
                                    prediction_val = pred_raw

                            def fmt(v):
                                if v is None:
                                    return "null"
                                if isinstance(v, (dict, list)):
                                    s = json.dumps(v, ensure_ascii=False, indent=2)
                                else:
                                    s = str(v)
                                return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", s)

                            rows.append({
                                "eval_res": fmt(eval_res),
                                "eval_details": fmt(eval_details),
                                "origin_prompt": fmt(origin_prompt),
                                "prediction": fmt(prediction_val),
                                "gold": fmt(gold),
                            })
                        except json.JSONDecodeError:
                            continue

                if rows:
                    out_name = os.path.splitext(os.path.basename(jf))[0] + ".xlsx"
                    out_path = os.path.join(task_out_dir, out_name)
                    pd.DataFrame(rows).to_excel(out_path, index=False)
                    print(f"  📄 {out_path}")

            # 拷贝 summary 子目录
            suite_summary_src = eval_dir / eval_version / "summary" / suite_name
            if suite_summary_src.exists():
                summary_dst = os.path.join(task_out_dir, "summary")
                if not os.path.exists(summary_dst):
                    shutil.copytree(str(suite_summary_src), summary_dst)
                    print(f"  📁 summary → {summary_dst}")


def main():
    parser = argparse.ArgumentParser(
        description="将单次评测结果导出为 Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--eval-dir",
        required=True,
        help="mixed_eval_* 根目录",
    )
    parser.add_argument(
        "--eval-version",
        default="eval_init",
        help="评测版本子目录名（默认 eval_init）",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="输出目录（默认 eval-dir/../benchmark/outputs/）",
    )
    args = parser.parse_args()

    output_dir = args.output_dir
    if output_dir is None:
        output_dir = str(Path(args.eval_dir).parent / "benchmark" / "outputs")

    process(
        eval_dir=args.eval_dir,
        eval_version=args.eval_version,
        output_base_dir=output_dir,
    )


if __name__ == "__main__":
    main()