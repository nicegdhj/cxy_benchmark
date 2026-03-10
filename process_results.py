import os
import shutil
import json
import pandas as pd
from datetime import datetime
import glob
from openpyxl.styles import PatternFill

def load_mappings(output_base_dir="outputs"):
    # 1. Task mapping
    task_mapping_file = os.path.join(output_base_dir, "评测任务文件名对应.xlsx")
    task_mapping = {}
    df_task_mapping = None
    if os.path.exists(task_mapping_file):
        df_task_mapping = pd.read_excel(task_mapping_file)
        if len(df_task_mapping.columns) >= 2:
            # Assuming col 0 is "任务文件名", col 1 is "任务数据集"
            for _, row in df_task_mapping.iterrows():
                k = str(row.iloc[0]).strip()
                v = str(row.iloc[1]).strip()
                if pd.notna(k) and pd.notna(v) and k != 'nan' and v != 'nan':
                    task_mapping[k] = v

    # 2. Experiment mapping
    exp_mapping_file = os.path.join(output_base_dir, "实验设置.xlsx")
    id_to_name = {}
    df_exp_mapping = None
    if os.path.exists(exp_mapping_file):
        df_exp_mapping = pd.read_excel(exp_mapping_file)
        id_col = '编号' if '编号' in df_exp_mapping.columns else df_exp_mapping.columns[0]
        # Drop rows where id_col is NaN
        df_exp_mapping = df_exp_mapping.dropna(subset=[id_col])
        set_col = df_exp_mapping.columns[1] # "实验设置"
        
        for _, row in df_exp_mapping.iterrows():
            try:
                row_id = int(float(row[id_col]))
                set_name = str(row[set_col]).strip()
                if set_name.endswith('.json'):
                    set_name = set_name[:-5]
                id_to_name[row_id] = set_name
            except (ValueError, TypeError):
                continue
    
    return task_mapping, df_task_mapping, id_to_name, df_exp_mapping


def get_mapped_exp_name(raw_name, id_to_name):
    if raw_name == 'baseline':
        return 'baseline'
    
    if raw_name.startswith('pt'):
        import re
        match = re.search(r'pt(\d+)_', raw_name)
        if match:
            pt_id = int(match.group(1))
            if pt_id in id_to_name:
                return id_to_name[pt_id]
    return raw_name

def process_results(base_fmt_dir="outputs/fmt", output_base_dir="outputs"):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target_dir = os.path.join(output_base_dir, f"aggregated_reports_{timestamp}")
    os.makedirs(target_dir, exist_ok=True)
    
    summary_excel_path = os.path.join(target_dir, f"总体汇总_{timestamp}.xlsx")
    
    task_mapping, df_task_mapping, id_to_name, df_exp_mapping = load_mappings(output_base_dir)
    
    if df_exp_mapping is not None:
        id_col = '编号' if '编号' in df_exp_mapping.columns else df_exp_mapping.columns[0]
        # Insert raw id mapping reference into df as "实验组编号"
        new_col = []
        for _, row in df_exp_mapping.iterrows():
             try:
                 new_col.append(f"pt{int(float(row[id_col]))}_sft0")
             except:
                 new_col.append("ptNaN_sft0")
        df_exp_mapping.insert(0, "实验组编号", new_col)

    all_reports = {}
    
    # 1. 扫描所有的 report.json
    for raw_exp_group in os.listdir(base_fmt_dir):
        group_path = os.path.join(base_fmt_dir, raw_exp_group)
        if not os.path.isdir(group_path):
            continue
        
        report_file = os.path.join(group_path, "report.json")
        if not os.path.exists(report_file):
            print(f"Warning: No report.json found in {group_path}")
            continue
            
        with open(report_file, 'r', encoding='utf-8') as f:
            try:
                data = json.load(f)
                mapped_exp_group = get_mapped_exp_name(raw_exp_group, id_to_name)
                all_reports[(raw_exp_group, mapped_exp_group)] = data
            except json.JSONDecodeError:
                print(f"Error reading {report_file}")
                
    if not all_reports:
        print("No valid reports found!")
        return

    # 准备总体对比 Sheet 
    all_raw_task_names = set()
    for _, report_data in all_reports.items():
        if "tasks" in report_data:
            for task in report_data["tasks"]:
                all_raw_task_names.add(task.get("task", "Unknown Task"))
                
    task_rows = []
    
    # Sort groups: baseline first
    exp_groups_sorted = list(all_reports.keys())
    baseline_tuple = next((t for t in exp_groups_sorted if t[0] == 'baseline'), None)
    if baseline_tuple:
        exp_groups_sorted.remove(baseline_tuple)
        exp_groups_sorted.insert(0, baseline_tuple)

    for raw_task_name in sorted(all_raw_task_names):
        mapped_task_name = task_mapping.get(raw_task_name, raw_task_name)
        row_dict = {"Task": mapped_task_name}
        for group_tuple in exp_groups_sorted:
            mapped_exp_group = group_tuple[1]
            report_data = all_reports[group_tuple]
            
            matching_task = next((t for t in report_data.get("tasks", []) if t.get("task") == raw_task_name), None)
            if matching_task:
                row_dict[mapped_exp_group] = matching_task.get("accuracy", None)
            else:
                row_dict[mapped_exp_group] = None
        task_rows.append(row_dict)
        
    df_overview = pd.DataFrame(task_rows)
    cols = ['Task'] + [t[1] for t in exp_groups_sorted]
    df_overview = df_overview[cols]

    if not df_overview.empty:
        df_overview.set_index("Task", inplace=True)
    
    # 生成 Excel
    with pd.ExcelWriter(summary_excel_path, engine='openpyxl') as writer:
        df_overview.to_excel(writer, sheet_name="总体对比")
        worksheet = writer.sheets["总体对比"]
        
        # Colors
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
        
        has_baseline = baseline_tuple is not None
        baseline_col_idx = 2 if has_baseline else None
        
        if has_baseline:
            for row in range(2, len(df_overview) + 2):
                baseline_val = worksheet.cell(row=row, column=baseline_col_idx).value
                # Color whole baseline
                worksheet.cell(row=row, column=baseline_col_idx).fill = yellow_fill
                
                if isinstance(baseline_val, (int, float)):
                    # Compare other columns with baseline
                    for col in range(baseline_col_idx + 1, len(cols) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        comp_val = cell.value
                        if isinstance(comp_val, (int, float)):
                            if comp_val < baseline_val:
                                cell.fill = red_fill
                            elif comp_val > baseline_val:
                                cell.fill = green_fill
        
        # Append mappings below the overview table
        current_row = len(df_overview) + 4
        
        if df_task_mapping is not None:
            worksheet.cell(row=current_row, column=1).value = "【附录 1】任务映射关系"
            df_task_mapping.to_excel(writer, sheet_name="总体对比", startrow=current_row, index=False)
            current_row += len(df_task_mapping) + 3
            
        if df_exp_mapping is not None:
            worksheet.cell(row=current_row, column=1).value = "【附录 2】实验组映射关系"
            df_exp_mapping.to_excel(writer, sheet_name="总体对比", startrow=current_row, index=False)
        
        # 逐个生成实验内部情况 Sheet
        for raw_exp_group, mapped_exp_group in all_reports.keys():
            report_data = all_reports[(raw_exp_group, mapped_exp_group)]
            
            summary = report_data.get("summary", {})
            sum_data = {
                "Type": ["Custom", "Generic"],
                "Count": [summary.get("custom", {}).get("count", 0), summary.get("generic", {}).get("count", 0)],
                "Total Duration (sec)": [summary.get("custom", {}).get("total_duration_sec", 0), summary.get("generic", {}).get("total_duration_sec", 0)],
                "Avg Accuracy": [summary.get("custom", {}).get("avg_accuracy", 0), summary.get("generic", {}).get("avg_accuracy", 0)]
            }
            df_sum = pd.DataFrame(sum_data)
            
            tasks_list = report_data.get("tasks", [])
            for task_dict in tasks_list:
                raw_task = task_dict.get("task", "unknown")
                mapped_task = task_mapping.get(raw_task, raw_task)
                
                task_dict["task"] = mapped_task
                task_dict["details_dir"] = f"{mapped_task}/{mapped_exp_group}"
                
            df_tasks = pd.DataFrame(tasks_list)
            
            sheet_name = str(mapped_exp_group)
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
                
            base_sheet_name = sheet_name
            counter = 1
            while sheet_name in writer.sheets:
                 sheet_name = f"{base_sheet_name[:28]}_{counter}"
                 counter += 1

            df_sum.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
            df_tasks.to_excel(writer, sheet_name=sheet_name, index=False, startrow=len(df_sum) + 3)

    print(f"Created overview excel at: {summary_excel_path}")
    
    # 2. 提取并转换明细 Excel
    for raw_exp_group, mapped_exp_group in all_reports.keys():
        
        group_path = os.path.join(base_fmt_dir, raw_exp_group)
        report_file = os.path.join(group_path, "report.json")
        try:
             with open(report_file, 'r', encoding='utf-8') as f:
                  raw_data = json.load(f)
        except Exception:
             continue
             
        for task in raw_data.get("tasks", []):
            raw_task_name = task.get("task", "unknown_task")
            mapped_task_name = task_mapping.get(raw_task_name, raw_task_name)
            details_dir_rel = task.get("details_dir")
            
            if not details_dir_rel:
                continue
                
            details_dir_abs = os.path.join(base_fmt_dir, raw_exp_group, details_dir_rel)
            results_dir = os.path.join(details_dir_abs, "results")
            
            task_out_dir = os.path.join(target_dir, mapped_task_name, mapped_exp_group)
            
            # Copy summary directory if it exists
            summary_dir_src = os.path.join(details_dir_abs, "summary")
            if os.path.exists(summary_dir_src) and os.path.isdir(summary_dir_src):
                summary_dir_dst = os.path.join(task_out_dir, "summary")
                if not os.path.exists(summary_dir_dst):
                    try:
                        shutil.copytree(summary_dir_src, summary_dir_dst)
                    except Exception as e:
                        print(f"Warning: Failed to copy summary dir for {raw_exp_group} - {mapped_task_name}: {e}")
            
            if not os.path.exists(results_dir):
                print(f"Warning: Results dir not found for {raw_exp_group} - {raw_task_name}: {results_dir}")
                continue
                
            jsonl_files = glob.glob(os.path.join(results_dir, "**", "*.jsonl"), recursive=True)
            if not jsonl_files:
                continue
                
            for jf in jsonl_files:
                rows = []
                with open(jf, 'r', encoding='utf-8') as f:
                    for line in f:
                        line = line.strip()
                        if not line: continue
                        try:
                            obj = json.loads(line)
                            
                            eval_res = obj.get("eval_res")
                            eval_details = obj.get("eval_details")
                            prediction_raw = obj.get("prediction")
                            
                            origin_prompt = None
                            prediction_val = None
                            gold = None
                            
                            if isinstance(prediction_raw, dict):
                                origin_prompt = prediction_raw.get("origin_prompt", "null")
                                prediction_val = prediction_raw.get("prediction", "null")
                                gold = prediction_raw.get("gold", "null")
                            elif isinstance(prediction_raw, str):
                                try:
                                    parsed_pred = json.loads(prediction_raw)
                                    if isinstance(parsed_pred, dict):
                                        origin_prompt = parsed_pred.get("origin_prompt", "null")
                                        prediction_val = parsed_pred.get("prediction", "null")
                                        gold = parsed_pred.get("gold", "null")
                                    else:
                                        prediction_val = prediction_raw
                                except json.JSONDecodeError:
                                    prediction_val = prediction_raw
                            
                            def format_field(f):
                                if f is None: return "null"
                                if isinstance(f, (dict, list)):
                                    return json.dumps(f, ensure_ascii=False, indent=2)
                                return str(f)
                            
                            rows.append({
                                "eval_res": format_field(eval_res),
                                "eval_details": format_field(eval_details),
                                "origin_prompt": format_field(origin_prompt),
                                "prediction": format_field(prediction_val),
                                "gold": format_field(gold)
                            })
                        except json.JSONDecodeError:
                            continue
                            
                if rows:
                    df = pd.DataFrame(rows)
                    
                    task_out_dir = os.path.join(target_dir, mapped_task_name, mapped_exp_group)
                    os.makedirs(task_out_dir, exist_ok=True)
                    
                    base_name = os.path.basename(jf)
                    out_name = os.path.splitext(base_name)[0] + ".xlsx"
                    out_path = os.path.join(task_out_dir, out_name)
                    
                    df.to_excel(out_path, index=False)
                    print(f"Created: {out_path}")

if __name__ == "__main__":
    process_results()
