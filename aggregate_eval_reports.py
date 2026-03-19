#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
aggregate_eval_reports.py - 汇总 eval_judge.py 产出的评测结果

读取 fmt/ 下所有实验组的 eval_{version}/report.json，
生成与 process_results.py 相同格式的汇总 Excel。

用法：
    python aggregate_eval_reports.py \\
        --fmt-dir /path/to/fmt \\
        --eval-version v3_reeval \\
        --output-dir /path/to/benchmark/outputs

    # 不传 --output-dir 则输出到 fmt-dir/../benchmark/outputs/
"""

import argparse
import glob
import json
import os
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill


def load_mappings(output_base_dir: str):
    """读取任务名映射和实验组映射（与 process_results.py 相同逻辑）。"""
    task_mapping = {}
    df_task_mapping = None
    task_mapping_file = os.path.join(output_base_dir, "评测任务文件名对应.xlsx")
    if os.path.exists(task_mapping_file):
        df_task_mapping = pd.read_excel(task_mapping_file)
        if len(df_task_mapping.columns) >= 2:
            for _, row in df_task_mapping.iterrows():
                k = str(row.iloc[0]).strip()
                v = str(row.iloc[1]).strip()
                if pd.notna(k) and pd.notna(v) and k != "nan" and v != "nan":
                    task_mapping[k] = v

    id_to_name = {}
    df_exp_mapping = None
    exp_mapping_file = os.path.join(output_base_dir, "实验设置.xlsx")
    if os.path.exists(exp_mapping_file):
        df_exp_mapping = pd.read_excel(exp_mapping_file)
        id_col = "编号" if "编号" in df_exp_mapping.columns else df_exp_mapping.columns[0]
        df_exp_mapping = df_exp_mapping.dropna(subset=[id_col])
        set_col = df_exp_mapping.columns[1]
        for _, row in df_exp_mapping.iterrows():
            try:
                row_id = int(float(row[id_col]))
                set_name = str(row[set_col]).strip()
                if set_name.endswith(".json"):
                    set_name = set_name[:-5]
                id_to_name[row_id] = set_name
            except (ValueError, TypeError):
                continue

    return task_mapping, df_task_mapping, id_to_name, df_exp_mapping


def get_mapped_suite(raw_suite: str, task_mapping: dict) -> str:
    """查映射表，找不到时尝试去掉 _suite 后缀再查。"""
    if raw_suite in task_mapping:
        return task_mapping[raw_suite]
    stripped = raw_suite.removesuffix("_suite")
    return task_mapping.get(stripped, raw_suite)


def get_mapped_exp_name(raw_name: str, id_to_name: dict) -> str:
    if raw_name in ("baseline", "baseline_nothink"):
        return raw_name
    if raw_name.startswith("pt"):
        import re
        # 匹配 pt{id}_sft0 或 pt{id}_sf0，并捕获其后的额外后缀（如 _exp0317）
        match = re.match(r"pt(\d+)_sf[t]?0(.*)", raw_name)
        if match:
            pt_id = int(match.group(1))
            extra = match.group(2)  # e.g. "_exp0317" 或 ""
            if pt_id in id_to_name:
                return id_to_name[pt_id] + extra
        # fallback：仅提取 id
        match2 = re.search(r"pt(\d+)_", raw_name)
        if match2:
            pt_id = int(match2.group(1))
            if pt_id in id_to_name:
                return id_to_name[pt_id]
    return raw_name


def find_eval_report(group_dir: Path, eval_version: str) -> Path | None:
    """找到 eval_{version}/report.json。"""
    report = group_dir / eval_version / "report.json"
    return report if report.exists() else None


def process(fmt_dir: str, eval_version: str, output_base_dir: str):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target_dir = os.path.join(output_base_dir, f"aggregated_reports_{timestamp}")
    os.makedirs(target_dir, exist_ok=True)
    summary_excel_path = os.path.join(target_dir, f"总体汇总_{timestamp}.xlsx")

    task_mapping, df_task_mapping, id_to_name, df_exp_mapping = load_mappings(output_base_dir)

    if df_exp_mapping is not None:
        id_col = "编号" if "编号" in df_exp_mapping.columns else df_exp_mapping.columns[0]
        df_exp_mapping = df_exp_mapping.dropna(subset=[id_col])
        new_col = []
        for _, row in df_exp_mapping.iterrows():
            try:
                new_col.append(f"pt{int(float(row[id_col]))}_sft0")
            except Exception:
                new_col.append("ptNaN_sft0")
        df_exp_mapping.insert(0, "实验组编号", new_col)

    # ── 扫描所有实验组 ────────────────────────────────────────────────────────
    all_reports = {}  # (raw_name, mapped_name) -> report data

    for entry in sorted(os.scandir(fmt_dir), key=lambda e: e.name):
        if not entry.is_dir():
            continue
        raw_name = entry.name
        group_dir = Path(entry.path)
        report_path = find_eval_report(group_dir, eval_version)
        if report_path is None:
            print(f"⚠️  跳过 {raw_name}：找不到 {eval_version}/report.json")
            continue

        try:
            data = json.loads(report_path.read_text(encoding="utf-8"))
            mapped_name = get_mapped_exp_name(raw_name, id_to_name)
            all_reports[(raw_name, mapped_name)] = data
        except Exception as e:
            print(f"❌ 读取 {report_path} 失败：{e}")

    if not all_reports:
        print("❌ 未找到任何有效 report.json，请检查 --eval-version 是否正确")
        sys.exit(1)

    print(f"✅ 找到 {len(all_reports)} 个实验组")

    # ── 准备总体对比表 ────────────────────────────────────────────────────────
    # 新 report.json: tasks 是 list，每项有 suite/task/accuracy 字段
    all_suites = set()
    for _, report_data in all_reports.items():
        for t in report_data.get("tasks", []):
            all_suites.add(t.get("suite", t.get("task", "unknown")))

    # baseline 排第一
    exp_groups_sorted = list(all_reports.keys())
    baseline_tuple = next((t for t in exp_groups_sorted if t[0] == "baseline"), None)
    if baseline_tuple:
        exp_groups_sorted.remove(baseline_tuple)
        exp_groups_sorted.insert(0, baseline_tuple)

    task_rows = []
    for raw_suite in sorted(all_suites):
        mapped_suite = get_mapped_suite(raw_suite, task_mapping)
        row_dict = {"Task": mapped_suite}
        for group_tuple in exp_groups_sorted:
            mapped_exp = group_tuple[1]
            report_data = all_reports[group_tuple]
            matching = next(
                (t for t in report_data.get("tasks", [])
                 if t.get("suite", t.get("task")) == raw_suite),
                None,
            )
            row_dict[mapped_exp] = matching.get("accuracy") if matching else None
        task_rows.append(row_dict)

    df_overview = pd.DataFrame(task_rows)
    cols = ["Task"] + [t[1] for t in exp_groups_sorted]
    df_overview = df_overview[cols]
    if not df_overview.empty:
        df_overview.set_index("Task", inplace=True)

    # ── 生成 Excel ───────────────────────────────────────────────────────────
    with pd.ExcelWriter(summary_excel_path, engine="openpyxl") as writer:
        df_overview.to_excel(writer, sheet_name="总体对比")
        worksheet = writer.sheets["总体对比"]

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill    = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        green_fill  = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

        has_baseline = baseline_tuple is not None
        if has_baseline:
            for row in range(2, len(df_overview) + 2):
                baseline_val = worksheet.cell(row=row, column=2).value
                worksheet.cell(row=row, column=2).fill = yellow_fill
                if isinstance(baseline_val, (int, float)):
                    for col in range(3, len(cols) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        if isinstance(cell.value, (int, float)):
                            cell.fill = green_fill if cell.value > baseline_val else red_fill

        current_row = len(df_overview) + 4
        if df_task_mapping is not None:
            worksheet.cell(row=current_row, column=1).value = "【附录 1】任务映射关系"
            df_task_mapping.to_excel(writer, sheet_name="总体对比", startrow=current_row, index=False)
            current_row += len(df_task_mapping) + 3
        if df_exp_mapping is not None:
            worksheet.cell(row=current_row, column=1).value = "【附录 2】实验组映射关系"
            df_exp_mapping.to_excel(writer, sheet_name="总体对比", startrow=current_row, index=False)

        # 每个实验组单独一个 Sheet
        for raw_name, mapped_name in all_reports.keys():
            report_data = all_reports[(raw_name, mapped_name)]
            summary = report_data.get("summary", {})
            sum_data = {
                "Type": ["Custom", "Generic"],
                "Count": [
                    summary.get("custom", {}).get("count", 0),
                    summary.get("generic", {}).get("count", 0),
                ],
                "Total Duration (sec)": [
                    summary.get("custom", {}).get("total_duration_sec", 0),
                    summary.get("generic", {}).get("total_duration_sec", 0),
                ],
                "Avg Accuracy": [
                    summary.get("custom", {}).get("avg_accuracy", 0),
                    summary.get("generic", {}).get("avg_accuracy", 0),
                ],
            }
            df_sum = pd.DataFrame(sum_data)

            tasks_list = []
            for t in report_data.get("tasks", []):
                raw_suite = t.get("suite", t.get("task", "unknown"))
                tasks_list.append({
                    "suite":       raw_suite,
                    "task":        get_mapped_suite(raw_suite, task_mapping),
                    "type":        t.get("type", "-"),
                    "eval_type":   t.get("eval_type", "-"),
                    "status":      t.get("status", "-"),
                    "accuracy":    t.get("accuracy"),
                    "num_samples": t.get("num_samples"),
                    "duration_sec": t.get("duration_sec"),
                })
            df_tasks = pd.DataFrame(tasks_list)

            sheet_name = str(mapped_name)[:31]
            base = sheet_name
            i = 1
            while sheet_name in writer.sheets:
                sheet_name = f"{base[:28]}_{i}"
                i += 1

            df_sum.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
            df_tasks.to_excel(writer, sheet_name=sheet_name, index=False, startrow=len(df_sum) + 3)

    print(f"📊 总体汇总 Excel: {summary_excel_path}")

    # ── 提取明细 JSONL → Excel ───────────────────────────────────────────────
    for raw_name, mapped_name in all_reports.keys():
        group_dir = Path(fmt_dir) / raw_name
        eval_results_dir = group_dir / eval_version / "results"
        if not eval_results_dir.exists():
            continue

        report_data = all_reports[(raw_name, mapped_name)]
        suite_set = {t.get("suite", t.get("task")) for t in report_data.get("tasks", [])}

        for suite in suite_set:
            mapped_suite = get_mapped_suite(suite, task_mapping)
            suite_dir = eval_results_dir / suite
            if not suite_dir.exists():
                # 兼容旧结构（results/ 下直接有 jsonl）
                suite_dir = eval_results_dir

            jsonl_files = glob.glob(str(suite_dir / "**" / "*.jsonl"), recursive=True)
            if not jsonl_files:
                # fallback: 在 results/ 下按 suite 名搜索
                jsonl_files = glob.glob(
                    str(eval_results_dir / "**" / f"*{suite}*_details.jsonl"), recursive=True
                )
            if not jsonl_files:
                continue

            task_out_dir = os.path.join(target_dir, mapped_suite, mapped_name)
            os.makedirs(task_out_dir, exist_ok=True)

            for jf in jsonl_files:
                rows = []
                with open(jf, encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        if not line:
                            continue
                        try:
                            obj = json.loads(line)
                            eval_res = obj.get("eval_res")
                            eval_details = obj.get("eval_details")
                            pred_raw = obj.get("prediction")
                            origin_prompt = prediction_val = gold = None

                            if isinstance(pred_raw, dict):
                                origin_prompt = pred_raw.get("origin_prompt", "null")
                                prediction_val = pred_raw.get("prediction", "null")
                                gold = pred_raw.get("gold", "null")
                            elif isinstance(pred_raw, str):
                                try:
                                    parsed = json.loads(pred_raw)
                                    if isinstance(parsed, dict):
                                        origin_prompt = parsed.get("origin_prompt", "null")
                                        prediction_val = parsed.get("prediction", "null")
                                        gold = parsed.get("gold", "null")
                                    else:
                                        prediction_val = pred_raw
                                except json.JSONDecodeError:
                                    prediction_val = pred_raw

                            def fmt(v):
                                if v is None: return "null"
                                if isinstance(v, (dict, list)):
                                    s = json.dumps(v, ensure_ascii=False, indent=2)
                                else:
                                    s = str(v)
                                # 过滤 Excel 非法控制字符（openpyxl 不接受 \x00-\x08 \x0b \x0c \x0e-\x1f）
                                return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", s)

                            rows.append({
                                "eval_res": fmt(eval_res),
                                "eval_details": fmt(eval_details),
                                "origin_prompt": fmt(origin_prompt),
                                "prediction": fmt(prediction_val),
                                "gold": fmt(gold),
                            })
                        except json.JSONDecodeError:
                            continue

                if rows:
                    out_name = os.path.splitext(os.path.basename(jf))[0] + ".xlsx"
                    out_path = os.path.join(task_out_dir, out_name)
                    pd.DataFrame(rows).to_excel(out_path, index=False)
                    print(f"  📄 {out_path}")

            # 拷贝 summary 子目录（对齐 process_results.py 的行为）
            suite_summary_src = group_dir / eval_version / "summary" / suite
            if suite_summary_src.exists():
                summary_dst = os.path.join(task_out_dir, "summary")
                if not os.path.exists(summary_dst):
                    shutil.copytree(str(suite_summary_src), summary_dst)
                    print(f"  📁 summary → {summary_dst}")

    print(f"\n✅ 汇总完成 → {target_dir}")


def main():
    parser = argparse.ArgumentParser(
        description="汇总 eval_judge.py 产出的评测结果为 Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--fmt-dir",
        default="/Users/jia/MyProjects/pythonProjects/cmcc_cxy/Bprocss/fmt",
        help="fmt 根目录（各实验组的父目录）",
    )
    parser.add_argument(
        "--eval-version",
        required=True,
        help="评测版本号，即 eval_judge.py 的 --eval-version 参数",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="汇总输出目录（默认 fmt-dir/../benchmark/outputs/）",
    )
    args = parser.parse_args()

    output_dir = args.output_dir
    if output_dir is None:
        output_dir = str(Path(args.fmt_dir).parent / "benchmark" / "outputs")

    process(
        fmt_dir=args.fmt_dir,
        eval_version=args.eval_version,
        output_base_dir=output_dir,
    )


if __name__ == "__main__":
    main()
